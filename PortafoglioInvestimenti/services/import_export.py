from pathlib import Path
import csv

try:
    import openpyxl
except ImportError:
    openpyxl = None


# =====================================================
# ESPORTAZIONE EXCEL
# =====================================================

def esporta_excel(strumenti, file_path):

    if openpyxl is None:
        raise ImportError(
            "openpyxl non installato.\n"
            "Installa con:\n"
            "pip install openpyxl"
        )

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Portafoglio"

    intestazioni = [
        "Ticker",
        "Nome",
        "Quantità",
        "Prezzo Medio",
        "Valuta",
        "Settore",
        "Note"
    ]

    for col, valore in enumerate(intestazioni, start=1):
        ws.cell(row=1, column=col).value = valore

    riga = 2

    for s in strumenti:

        ws.cell(riga, 1).value = s["ticker"]
        ws.cell(riga, 2).value = s["nome"]
        ws.cell(riga, 3).value = s["quantita"]
        ws.cell(riga, 4).value = s["prezzo_medio"]
        ws.cell(riga, 5).value = s["valuta"]
        ws.cell(riga, 6).value = s["settore"]
        ws.cell(riga, 7).value = s["note"]

        riga += 1

    for colonna in ws.columns:

        lunghezza = 0

        lettera = colonna[0].column_letter

        for cella in colonna:

            try:
                lunghezza = max(
                    lunghezza,
                    len(str(cella.value))
                )
            except Exception:
                pass

        ws.column_dimensions[lettera].width = lunghezza + 3

    wb.save(file_path)


# =====================================================
# ESPORTAZIONE CSV
# =====================================================

def esporta_csv(strumenti, file_path):

    with open(
        file_path,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as file:

        writer = csv.writer(file, delimiter=";")

        writer.writerow([
            "Ticker",
            "Nome",
            "Quantità",
            "Prezzo Medio",
            "Valuta",
            "Settore",
            "Note"
        ])

        for s in strumenti:

            writer.writerow([

                s["ticker"],

                s["nome"],

                s["quantita"],

                s["prezzo_medio"],

                s["valuta"],

                s["settore"],

                s["note"]

            ])


# =====================================================
# IMPORTAZIONE CSV
# =====================================================

def importa_csv(file_path):

    strumenti = []

    with open(
        file_path,
        newline="",
        encoding="utf-8-sig"
    ) as file:

        reader = csv.DictReader(file, delimiter=";")

        for riga in reader:

            strumenti.append({

                "ticker": riga["Ticker"],

                "nome": riga["Nome"],

                "quantita": float(
                    str(riga["Quantità"]).replace(",", ".")
                ),

                "prezzo": float(
                    str(riga["Prezzo Medio"]).replace(",", ".")
                ),

                "valuta": riga["Valuta"],

                "settore": riga.get("Settore", ""),

                "note": riga.get("Note", "")

            })

    return strumenti


# =====================================================
# IMPORTAZIONE EXCEL
# =====================================================

def importa_excel(file_path):

    if openpyxl is None:
        raise ImportError(
            "openpyxl non installato."
        )

    wb = openpyxl.load_workbook(file_path)

    ws = wb.active

    strumenti = []

    for riga in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if riga[0] is None:
            continue

        strumenti.append({

            "ticker": str(riga[0]),

            "nome": str(riga[1]),

            "quantita": float(riga[2]),

            "prezzo": float(riga[3]),

            "valuta": str(riga[4]),

            "settore": "" if riga[5] is None else str(riga[5]),

            "note": "" if riga[6] is None else str(riga[6])

        })

    return strumenti


# =====================================================
# DIRECTORY PREDEFINITA
# =====================================================

def cartella_export():

    cartella = Path.home() / "Documents"

    cartella.mkdir(exist_ok=True)

    return cartella